# -*- coding: utf-8 -*-
import os
import re
import io
import zipfile
import tempfile
import datetime
import gc 
import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from scipy.stats import ttest_ind
from openpyxl import load_workbook
import streamlit as st
import plotly.graph_objects as go

# =====================================================
# Core Logic (Matplotlib saving & Math)
# =====================================================
def parse_phases_from_wave(xls):
    for s in xls.sheet_names:
        if "assay" in s.lower():
            df = pd.read_excel(xls, sheet_name=s, header=None)
            target_aliases = {
                "Baseline": ["baseline"], "Etomoxir": ["eto"],
                "Oligo": ["oligo"], "FCCP": ["fccp"], "Rot": ["rot", "rra"]
            }
            phase_col_map = {}
            found_phases = []
            for i, row in df.iterrows():
                row_strs = [str(x).lower().strip() for x in row.values]
                if "baseline" in row_strs:
                    for col_idx, cell_val in enumerate(row_strs):
                        for std_name, aliases in target_aliases.items():
                            if std_name not in found_phases:
                                if any(a in cell_val for a in aliases):
                                    phase_col_map[std_name] = col_idx
                                    found_phases.append(std_name)
                                    break
                if phase_col_map and any("cycles" in str(x).lower() for x in row_strs):
                    parsed_any = False
                    result = {}
                    for p, col_idx in phase_col_map.items():
                        if col_idx < len(row.values):
                            cell_val = str(row.values[col_idx]).lower()
                            match = re.search(r'\d+', cell_val)
                            if match:
                                result[p] = int(match.group(0))
                                parsed_any = True
                    if parsed_any:
                        return [{"name": p, "cycles": result.get(p, 3)} for p in found_phases]
    return [{"name": "Baseline", "cycles": 3}, {"name": "Oligo", "cycles": 3},
            {"name": "FCCP", "cycles": 3}, {"name": "Rot", "cycles": 3}]

def derive_ranges_from_cycles(cycle_dict, phase_order):
    cur = 1
    ranges = {}
    for k in phase_order:
        if k in cycle_dict:
            ranges[k] = list(range(cur, cur + cycle_dict[k]))
            cur += cycle_dict[k]
    return ranges

def parse_cell_counts(xls):
    for s in xls.sheet_names:
        if "assay" in s.lower():
            df = pd.read_excel(xls, sheet_name=s, header=None)
            start_row, start_col = None, None
            for i, row in df.iterrows():
                row_strs = [str(x).lower().strip() for x in row.values]
                if "normalization values:" in row_strs:
                    start_row = i + 1  
                    for j, val in enumerate(df.iloc[i+1].values):
                        if str(val).strip().upper() == 'A':
                            start_col = j + 1  
                            break
                    break
            if start_row is None or start_col is None:
                start_row, start_col = 84, 2
            counts = {}
            rows_letters = "ABCDEFGH"
            for i, r_letter in enumerate(rows_letters):
                for j in range(12):
                    well = f"{r_letter}{j+1:02d}"
                    try:
                        val = df.iloc[start_row + i, start_col + j]
                        if pd.notna(val):
                            counts[well] = float(val)
                    except:
                        pass
            return counts
    return {}

def detect_wells_by_fg_theme(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    sheet = next(s for s in wb.sheetnames if "assay" in s.lower())
    ws = wb[sheet]
    selected, unselected = set(), set()
    start_row, start_col = None, None
    for r in range(1, 30):
        for c in range(1, 10):
            cell_val = str(ws.cell(row=r, column=c).value).strip()
            if cell_val in ["A01", "A1"]:
                start_row, start_col = r, c
                break
        if start_row: break
    if not start_row: start_row, start_col = 12, 3 
    rows = "ABCDEFGH"
    for r_idx, r in enumerate(rows):
        for c_idx in range(12):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            well = f"{r}{c_idx + 1:02d}"
            fg = cell.fill.fgColor if cell.fill else None
            if fg and fg.type == "theme" and fg.theme == 0:
                unselected.add(well)
            else:
                selected.add(well)
    return selected, unselected

def load_rate_sheet(xls, sheet_name, selected_wells):
    df = pd.read_excel(xls, sheet_name=sheet_name)
    df.columns = df.columns.str.lower()
    df = df[df["well"].isin(selected_wells)]
    df = df[~df["group"].isin(["Background", "Unassigned", "background", "unassigned"])]
    return df

def compute_metrics(df, ranges):
    df = df[~df["group"].isin(["Background", "Unassigned", "background", "unassigned"])].copy()
    last_base = max(ranges.get("Baseline", [1]))
    
    basal = df[df["measurement"] == last_base].groupby(["group", "well"], observed=True)["ocr"].mean()
    metrics = {}
    if "Oligo" in ranges:
        oligo = df[df["measurement"].isin(ranges["Oligo"])].groupby(["group", "well"], observed=True)["ocr"].min()
        metrics["ATP-linked Respiration"] = basal - oligo
    if "Rot" in ranges:
        rot = df[df["measurement"].isin(ranges["Rot"])].groupby(["group", "well"], observed=True)["ocr"].min()
        metrics["Basal Respiration"] = basal - rot
    if "Oligo" in ranges and "Rot" in ranges:
        metrics["Proton Leak"] = oligo - rot
    if "FCCP" in ranges and "Rot" in ranges:
        fccp = df[df["measurement"].isin(ranges["FCCP"])].groupby(["group", "well"], observed=True)["ocr"].max()
        metrics["Max Respiration"] = fccp - rot
    if "Etomoxir" in ranges:
        eto = df[df["measurement"].isin(ranges["Etomoxir"])].groupby(["group", "well"], observed=True)["ocr"].min()
        metrics["Fatty Acid Oxidation"] = basal - eto
        
    res = pd.DataFrame(metrics).reset_index().dropna(how='all', subset=list(metrics.keys()))
    res = res[~res["group"].isin(["Background", "Unassigned", "background", "unassigned"])]
    return res

def save_plate_qc(unselected, outdir):
    plate = np.zeros((8, 12))
    for r in range(8):
        for c in range(12):
            if f"{chr(ord('A') + r)}{c + 1:02d}" in unselected:
                plate[r, c] = 1
    fig, ax = plt.subplots(figsize=(8, 3))
    ax.imshow(plate, cmap="gray_r")
    ax.set_xticks(range(12)); ax.set_xticklabels(range(1, 13))
    ax.set_yticks(range(8)); ax.set_yticklabels(list("ABCDEFGH"))
    ax.set_title("Plate QC (Grey = Excluded)")
    fig.tight_layout()
    fig.savefig(os.path.join(outdir, "plate_qc.png"), dpi=300)
    plt.close(fig)

def _row_letters_to_index(row_letters: str) -> int:
    idx = 0
    for ch in row_letters: idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx

def _sanitize_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", str(s)).strip("_")

def export_prism_csvs(summary_df, outdir: str, suffix: str):
    df = summary_df.copy()
    df.columns = [c.strip() for c in df.columns]
    metric_cols = [c for c in df.columns if c not in ["group", "well"]]
    if not metric_cols: return
    row_letters = df["well"].astype(str).str.extract(r"^([A-Za-z]+)")[0].str.upper()
    col_nums = df["well"].astype(str).str.extract(r"(\d+)$")[0]
    df["_row_idx"] = row_letters.apply(_row_letters_to_index)
    df["_col_num"] = col_nums.astype(int)
    prism_dir = os.path.join(outdir, "prism_csv")
    os.makedirs(prism_dir, exist_ok=True)
    for metric in metric_cols:
        tmp = df[["group", "well", "_row_idx", "_col_num", metric]].copy()
        tmp = tmp.sort_values(["group", "_row_idx", "_col_num", "well"], kind="mergesort")
        tmp["_rep"] = tmp.groupby("group", observed=True).cumcount() + 1
        wide = tmp.pivot(index="_rep", columns="group", values=metric).sort_index()
        wide = wide.rename_axis(None, axis=0).rename_axis(None, axis=1)
        wide.to_csv(os.path.join(prism_dir, f"Prism_{_sanitize_filename(metric)}_{suffix}.csv"), index=False, encoding="utf-8-sig")

def save_results(df, outdir, suffix, control):
    os.makedirs(outdir, exist_ok=True)
    df.to_csv(os.path.join(outdir, f"summary_{suffix}.csv"), index=False, encoding="utf-8-sig")
    export_prism_csvs(df, outdir, suffix)

    for metric in df.columns[2:]:
        fig, ax = plt.subplots(figsize=(8, 5))
        g = df.groupby("group", observed=True)[metric]
        m, s = g.mean(), g.std()
        ax.bar(m.index, m.values, yerr=s.values, capsize=4, zorder=1, alpha=0.8, color="#4A90E2")
        ax.set_title(metric, fontweight="bold")

        for i, grp in enumerate(m.index):
            y_vals = df[df["group"] == grp][metric].dropna()
            if len(y_vals) > 0:
                x_vals = np.random.normal(i, 0.06, size=len(y_vals))
                ax.scatter(x_vals, y_vals, color='black', alpha=0.6, s=25, zorder=3)

        ylabel = "Cell Count" if metric == "Cell Number" else "OCR (pmol/min)" if suffix == "raw" else "OCR (pmol/min/Norm. Unit)"
        ax.set_ylabel(ylabel)

        if control in m.index:
            ctrl_vals = df[df["group"] == control][metric].dropna()
            valid_maxes = [df[metric].max()]
            if not m.isna().all() and not s.isna().all(): valid_maxes.append((m + s).max())
            global_ymax = max([v for v in valid_maxes if not np.isnan(v)] or [0])
            for i, grp in enumerate(m.index):
                if grp == control: continue
                vals = df[df["group"] == grp][metric].dropna()
                if len(ctrl_vals) >= 2 and len(vals) >= 2:
                    _, p = ttest_ind(ctrl_vals, vals, equal_var=False)
                    stars = "**" if p < 0.01 else "*" if p < 0.05 else "ns"
                    p_str = "p<0.0001" if p < 0.0001 else f"p={round(p, 4)}"
                    grp_max_val = max(m[grp] + s[grp], vals.max())
                    ax.text(i, grp_max_val + global_ymax * 0.03, f"{stars}\n{p_str}", ha="center", va="bottom", fontsize=8, linespacing=1.2)

        plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")
        y_min = df[metric].min()
        if not np.isnan(y_min):
            ax.set_ylim(min(0, y_min * 1.1) if y_min < 0 else 0, ax.get_ylim()[1] * 1.25)

        fig.tight_layout()
        fig.savefig(os.path.join(outdir, f"{metric}_{suffix}.png"), dpi=300)
        plt.close(fig)

def save_kinetic_graphs(df, outdir, suffix, ranges, phase_order, metric_col="ocr"):
    if metric_col not in df.columns: return
    g = df.groupby(["group", "measurement"], observed=True)[metric_col]
    m, s = g.mean().unstack(level=0), g.sem().unstack(level=0)
    if m.empty: return

    fig, ax = plt.subplots(figsize=(10, 5)) # Slightly wider for legend
    for col in m.columns:
        ax.errorbar(m.index, m[col], yerr=s[col], label=col, marker='o', capsize=3, markersize=5, linewidth=2)

    ymin, ymax = ax.get_ylim()
    for i in range(1, len(phase_order)):
        phase = phase_order[i]
        if phase in ranges and len(ranges[phase]) > 0:
            line_x = ranges[phase][0] - 0.5
            ax.axvline(x=line_x, color='black', linestyle='--', alpha=0.5, linewidth=1.5)
            ax.text(line_x, ymax * 0.95, phase, rotation=0, ha='center', va='top', fontsize=9, fontweight='bold',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.9))

    metric_upper = metric_col.upper()
    ax.set_title(f"Kinetic {metric_upper} Profile", fontsize=14, fontweight='bold', pad=15)
    ax.set_xlabel("Measurement", fontsize=12)
    ylabel_unit = "(pmol/min)" if metric_col == "ocr" else "(mpH/min)"
    ax.set_ylabel(f"{metric_upper} {ylabel_unit}" if suffix == "raw" else f"{metric_upper} {ylabel_unit}/Norm. Unit")
    
    # Legend outside to the right
    ax.legend(bbox_to_anchor=(1.02, 1), loc='upper left', frameon=True)
    ax.grid(True, linestyle=':', alpha=0.6)
    ax.set_ylim(bottom=min(0, ymin), top=ymax * 1.1)
    
    fig.tight_layout()
    os.makedirs(outdir, exist_ok=True)
    fig.savefig(os.path.join(outdir, f"Kinetic_{metric_upper}_{suffix}.png"), dpi=300, bbox_inches="tight")
    plt.close(fig)


# =====================================================
# Interactive Plotly Generation (Frontend)
# =====================================================
def create_plotly_kinetic(df, suffix, ranges, phase_order, metric_col="ocr"):
    if metric_col not in df.columns: return None
    g = df.groupby(["group", "measurement"], observed=True)[metric_col]
    m, s = g.mean().unstack(level=0), g.sem().unstack(level=0)
    if m.empty: return None

    fig = go.Figure()
    for col in m.columns:
        fig.add_trace(go.Scatter(
            x=m.index, y=m[col],
            error_y=dict(type='data', array=s[col], visible=True),
            mode='lines+markers', name=col,
            line=dict(width=2), marker=dict(size=6),
            hovertemplate="Mea %{x}<br>Mean: %{y:.2f}<br>SEM: %{error_y.array:.2f}<extra></extra>"
        ))
    
    for i in range(1, len(phase_order)):
        phase = phase_order[i]
        if phase in ranges and len(ranges[phase]) > 0:
            line_x = ranges[phase][0] - 0.5
            fig.add_vline(x=line_x, line_dash="dash", line_color="gray", opacity=0.7)
            fig.add_annotation(x=line_x, y=1.05, yref="paper", text=f"<b>{phase}</b>", showarrow=False, 
                               bgcolor="white", bordercolor="gray", borderwidth=1)
    
    metric_upper = metric_col.upper()
    ylabel = f"{metric_upper} " + ("(pmol/min)" if metric_col == "ocr" else "(mpH/min)")
    if suffix == "norm": ylabel += "/Norm. Unit"
    
    fig.update_layout(
        title=f"<b>Interactive Kinetic {metric_upper}</b>",
        xaxis_title="Measurement", yaxis_title=ylabel,
        hovermode="x unified", template="plotly_white",
        # 【修复】：将图例移到右侧垂直排列，X轴开启自动边距
        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02),
        xaxis=dict(automargin=True)
    )
    return fig

def create_plotly_bar(df, metric, suffix, control):
    g = df.groupby("group", observed=True)[metric]
    m, s = g.mean(), g.std()
    
    fig = go.Figure()
    
    # 【修复】：强制底层 X 轴为纯数字以解决散点错位问题
    x_positions = list(range(len(m.index)))
    
    fig.add_trace(go.Bar(
        name='Mean', x=x_positions, y=m.values,
        error_y=dict(type='data', array=s.values),
        marker_color="#4A90E2", opacity=0.7, showlegend=False
    ))
    
    if control in m.index:
        ctrl_vals = df[df["group"] == control][metric].dropna()
    else:
        ctrl_vals = []
        
    global_ymax = (m+s).max() if not (m.isna().all() or s.isna().all()) else df[metric].max()
    
    for i, grp in enumerate(m.index):
        y_vals = df[df["group"] == grp][metric].dropna()
        if len(y_vals) > 0:
            x_vals = np.random.normal(i, 0.05, size=len(y_vals))
            fig.add_trace(go.Scatter(
                x=x_vals, y=y_vals, mode='markers',
                marker=dict(color='black', size=5, opacity=0.6),
                name=grp, hoverinfo='y+name', showlegend=False
            ))
        
        if grp != control and len(ctrl_vals) >= 2 and len(y_vals) >= 2:
            _, p = ttest_ind(ctrl_vals, y_vals, equal_var=False)
            stars = "**" if p < 0.01 else "*" if p < 0.05 else "ns"
            p_str = "p<0.0001" if p < 0.0001 else f"p={round(p,4)}"
            grp_max = max(m[grp] + s[grp], y_vals.max())
            
            fig.add_annotation(
                x=i, y=grp_max + global_ymax * 0.05,
                text=f"<b>{stars}</b><br>{p_str}", showarrow=False,
                font=dict(size=11)
            )
            
    ylabel = "Cell Count" if metric == "Cell Number" else "OCR (pmol/min)" if suffix=="raw" else "OCR (pmol/min/Norm. Unit)"
    fig.update_layout(
        title=f"<b>{metric}</b>", yaxis_title=ylabel,
        template="plotly_white", 
        height=550, 
        # 【修复】：表面标签替换回真实名字，倾斜45度并自动扩展底边距防遮挡
        xaxis=dict(
            tickmode='array',
            tickvals=x_positions,
            ticktext=m.index,
            tickangle=-45,
            automargin=True
        ),
        margin=dict(b=120)
    )
    return fig

def create_plotly_plate_qc(unselected):
    z = np.zeros((8, 12))
    for r in range(8):
        for c in range(12):
            if f"{chr(ord('A')+r)}{c+1:02d}" in unselected:
                z[r, c] = 1
    fig = go.Figure(data=go.Heatmap(
        z=z[::-1], x=list(range(1, 13)), y=list("ABCDEFGH")[::-1],
        colorscale=[[0, "#E8F0FE"], [1, "#5F6368"]], showscale=False, xgap=2, ygap=2,
        hovertemplate="Well: %{y}%{x:02d}<extra></extra>"
    ))
    fig.update_layout(title="<b>Plate QC (Grey = Excluded)</b>", height=350, template="plotly_white",
                      xaxis=dict(side="top")) 
    return fig

# =====================================================
# Extractor & Caching Engine
# =====================================================
@st.cache_resource(show_spinner=False)
def extract_all_data(file_bytes):
    file_io = io.BytesIO(file_bytes)
    selected, unselected = detect_wells_by_fg_theme(file_io)
    file_io.seek(0)
    
    try:
        xls = pd.ExcelFile(file_io, engine='calamine')
    except Exception:
        file_io.seek(0)
        xls = pd.ExcelFile(file_io)
        
    phases = parse_phases_from_wave(xls)
    cell_counts = parse_cell_counts(xls)
    
    rate_raw = None
    if any("rate" == str(s).lower() for s in xls.sheet_names):
        actual_sheet = next(s for s in xls.sheet_names if str(s).lower() == "rate")
        rate_raw = load_rate_sheet(xls, actual_sheet, selected)
            
    rate_norm = None
    if any("normalized rate" == str(s).lower() for s in xls.sheet_names):
        actual_sheet = next(s for s in xls.sheet_names if str(s).lower() == "normalized rate")
        rate_norm = load_rate_sheet(xls, actual_sheet, selected)
            
    return unselected, phases, cell_counts, rate_raw, rate_norm

# =====================================================
# 工具函数：强制应用用户的自定义排版顺序
# =====================================================
def apply_custom_order(df, order):
    if df is None or df.empty: return df
    df = df[df["group"].isin(order)].copy()
    df["group"] = pd.Categorical(df["group"], categories=order, ordered=True)
    return df


# =====================================================
# Streamlit Web App UI
# =====================================================
st.set_page_config(page_title="Seahorse Web Tool", layout="wide", page_icon="🧬")

st.title("🧬 Seahorse Analysis Web Tool (Ultimate Edition)")
st.markdown("Powered by **Calamine** engine & **Plotly** Interactive Graphics.")

with st.sidebar:
    st.header("1. Upload Data")
    uploaded_file = st.file_uploader("Upload Wave Excel (.xlsx)", type=["xlsx"])
    
    if uploaded_file:
        file_bytes = uploaded_file.getvalue()
        
        with st.spinner("⚡ Extracting Data (Superfast Mode)..."):
            unselected, phases, cell_counts, rate_raw, rate_norm = extract_all_data(file_bytes)
        
        phase_order = [p["name"] for p in phases]
        defaults = {p["name"]: p["cycles"] for p in phases}
        
        st.header("2. Assay Settings")
        st.markdown("Cycles per Phase:")
        cycle_vars = {}
        cols = st.columns(2)
        for i, phase in enumerate(phase_order):
            with cols[i % 2]:
                cycle_vars[phase] = st.number_input(f"{phase}", min_value=1, max_value=20, value=defaults.get(phase, 3))
        
        st.markdown("---")
        
        st.header("3. X-Axis Layout & Control")
        st.info("💡 **Drag and drop** the tags below to customize the order of bars on your charts!")
        
        original_groups = []
        if rate_raw is not None:
            for g in rate_raw["group"].dropna():
                if g not in original_groups and g.lower() not in ["background", "unassigned"]:
                    original_groups.append(g)
        
        final_group_order = st.multiselect(
            "Chart X-Axis Order:", 
            options=original_groups, 
            default=original_groups
        )
        
        control_group = st.selectbox("Select Control Group (for P-values):", final_group_order) if final_group_order else ""

        run_btn = st.button("🚀 Run Analysis", type="primary", use_container_width=True)

if uploaded_file and run_btn:
    if not final_group_order:
        st.error("Please select at least one group in the sidebar to plot.")
        st.stop()

    with st.spinner("Processing Metrics & Rendering Plotly..."):
        temp_dir = tempfile.mkdtemp()
        save_plate_qc(unselected, temp_dir)
        
        with open(os.path.join(temp_dir, "analysis_info.txt"), "w", encoding="utf-8") as f:
            f.write("Seahorse Analysis Summary\n=========================\n\nExcluded Wells:\n")
            if unselected:
                for w in sorted(list(unselected), key=lambda x: (x[0], int(x[1:]))): f.write(f" - {w}\n")
            else:
                f.write(" - None\n")
            f.write("\n* Cell Count analysis bypasses the 'Unselected Well' filter.\n")
        
        if rate_raw is not None:
            rate_raw = apply_custom_order(rate_raw, final_group_order)
        if rate_norm is not None:
            rate_norm = apply_custom_order(rate_norm, final_group_order)
            
        cell_df = None
        if cell_counts and rate_raw is not None:
            cell_df = rate_raw[['group', 'well']].drop_duplicates()
            cell_df["Cell Number"] = cell_df["well"].map(cell_counts)
            cell_df = cell_df.dropna(subset=["Cell Number"])
        
        ranges = derive_ranges_from_cycles(cycle_vars, phase_order)
        
        st.success("✅ Analysis Complete! Layout adjusted perfectly. You can interact with charts below.")

        zip_buffer = io.BytesIO()
        
        # --- RAW DATA ---
        if rate_raw is not None:
            res_raw = compute_metrics(rate_raw, ranges)
            if cell_df is not None and not cell_df.empty:
                res_raw = pd.merge(res_raw, cell_df, on=["group", "well"], how="outer")
                
            res_raw = apply_custom_order(res_raw, final_group_order)
            res_raw = res_raw.sort_values(["group", "well"])
            
            # Backend Save
            save_results(res_raw, os.path.join(temp_dir, "raw"), "raw", control_group)
            save_kinetic_graphs(rate_raw, os.path.join(temp_dir, "raw"), "raw", ranges, phase_order, "ocr")
            save_kinetic_graphs(rate_raw, os.path.join(temp_dir, "raw"), "raw", ranges, phase_order, "ecar")
            
            # Frontend Render
            st.markdown("### 🔹 Raw Data Results")
            tab_k_raw, tab_b_raw = st.tabs(["📉 Kinetic Curves", "📊 Bar Charts"])
            
            with tab_k_raw:
                c1, c2 = st.columns(2)
                with c1:
                    fig_k_ocr = create_plotly_kinetic(rate_raw, "raw", ranges, phase_order, "ocr")
                    if fig_k_ocr: st.plotly_chart(fig_k_ocr, use_container_width=True)
                with c2:
                    fig_k_ecar = create_plotly_kinetic(rate_raw, "raw", ranges, phase_order, "ecar")
                    if fig_k_ecar: st.plotly_chart(fig_k_ecar, use_container_width=True)
                    
            with tab_b_raw:
                metrics = [c for c in res_raw.columns if c not in ["group", "well"]]
                cols = st.columns(3)
                for i, m in enumerate(metrics):
                    fig_bar = create_plotly_bar(res_raw, m, "raw", control_group)
                    with cols[i % 3]:
                        st.plotly_chart(fig_bar, use_container_width=True)

        # --- NORMALIZED DATA ---
        if rate_norm is not None:
            st.divider()
            res_norm = compute_metrics(rate_norm, ranges)
            if cell_df is not None and not cell_df.empty:
                res_norm = pd.merge(res_norm, cell_df, on=["group", "well"], how="outer")
            
            res_norm = apply_custom_order(res_norm, final_group_order)
            res_norm = res_norm.sort_values(["group", "well"])
            
            # Backend Save
            save_results(res_norm, os.path.join(temp_dir, "normalized"), "norm", control_group)
            save_kinetic_graphs(rate_norm, os.path.join(temp_dir, "normalized"), "norm", ranges, phase_order, "ocr")
            save_kinetic_graphs(rate_norm, os.path.join(temp_dir, "normalized"), "norm", ranges, phase_order, "ecar")

            # Frontend Render
            st.markdown("### 🔹 Normalized Data Results")
            tab_k_norm, tab_b_norm = st.tabs(["📉 Kinetic Curves", "📊 Bar Charts"])
            
            with tab_k_norm:
                c1, c2 = st.columns(2)
                with c1:
                    fig_k_ocr_n = create_plotly_kinetic(rate_norm, "norm", ranges, phase_order, "ocr")
                    if fig_k_ocr_n: st.plotly_chart(fig_k_ocr_n, use_container_width=True)
                with c2:
                    fig_k_ecar_n = create_plotly_kinetic(rate_norm, "norm", ranges, phase_order, "ecar")
                    if fig_k_ecar_n: st.plotly_chart(fig_k_ecar_n, use_container_width=True)
                    
            with tab_b_norm:
                metrics = [c for c in res_norm.columns if c not in ["group", "well"]]
                cols = st.columns(3)
                for i, m in enumerate(metrics):
                    fig_bar_n = create_plotly_bar(res_norm, m, "norm", control_group)
                    with cols[i % 3]:
                        st.plotly_chart(fig_bar_n, use_container_width=True)

        st.divider()
        st.markdown("### 🧫 Plate QC")
        st.plotly_chart(create_plotly_plate_qc(unselected), use_container_width=False)

        # Create ZIP
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)

        # Download Button
        st.sidebar.markdown("---")
        st.sidebar.download_button(
            label="📦 Download HD Matplotlib ZIP",
            data=zip_buffer.getvalue(),
            file_name=f"Seahorse_Results_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mime="application/zip",
            type="primary"
        )
        
        # 强制垃圾回收，保障云端不爆内存
        del rate_raw, rate_norm
        gc.collect()

elif not uploaded_file:
    st.info("👈 Please upload your Seahorse Wave Excel file in the sidebar to begin.")
